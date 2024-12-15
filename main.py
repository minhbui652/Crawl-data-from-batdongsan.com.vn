import undetected_chromedriver as uc
import time
import re
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.by import By
from datetime import date
import unicodedata

start_time = time.time()

def convert_price(price, area):
    if "tỷ" in price:
        # Loại bỏ dấu phẩy nhưng giữ nguyên dấu chấm thập phân
        price_value = float(re.sub(r'[^\d.]', '', price.replace(",", ".").replace("tỷ", "").strip())) * 1_000_000
        print(price, price_value)
        return price_value
    elif "triệu/m" in price:
        price_per_sqm = float(re.sub(r'[^\d.]', '', price.replace(",", ".").replace("triệu/m²", "").strip())) * 1_000
        price_value = price_per_sqm * area
        print(price, price_value)
    elif "Thỏa thuận" in price:
        price_value = None
    else:
        price_value = None

    return price_value

def extract_frontage_from_description(description):
    match = re.search(r"(?:mặt tiền|Mặt tiền)[:\s]*([\d.,]+)\s*m", description, re.IGNORECASE)
    if match:
        return match.group(1)
    return None

def extract_main_orientation(description):
    # Định nghĩa các từ chỉ hướng cần tìm
    directions = ["nam", "tây nam", "bắc", "đông bắc", "đông", "đông nam", "tây", "tây bắc"]
    
    # Đưa toàn bộ description về chữ thường để so khớp không phân biệt hoa thường
    description = description.lower()
    
    # Tạo hai pattern regex để tìm lần lượt "hướng chính" hoặc "hướng"
    main_pattern = r"hướng\s+chính(?:\s+\w+){0,3}\s+(" + "|".join(directions) + ")"
    general_pattern = r"hướng(?:\s+\w+){0,3}\s+(" + "|".join(directions) + ")"
    
    # Tìm "hướng chính" trước
    main_match = re.search(main_pattern, description)
    if main_match:
        orientation = main_match.group(1)
    else:
        # Nếu không có "hướng chính", tìm hướng đầu tiên
        general_match = re.search(general_pattern, description)
        if general_match:
            orientation = general_match.group(1)
        else:
            return None  # Trả về None nếu không tìm thấy hướng
    
    # Định dạng lại kết quả thành dạng "Tây - Nam", "Đông - Bắc", v.v.
    formatted_orientation = " - ".join(word.capitalize() for word in orientation.split())
    return formatted_orientation

def convert_comma_to_dot(value):
    # Kiểm tra nếu value không phải là None và là chuỗi
    if value and isinstance(value, str):
        # Thay thế dấu phẩy bằng dấu chấm
        return value.replace(",", ".")
    return value

def normalize_string(value):
    if isinstance(value, str):
        return unicodedata.normalize('NFC', value)
    return value

def remove_illegal_characters(text):
    # Loại bỏ các ký tự không hợp lệ bằng cách dùng biểu thức chính quy
    if isinstance(text, str):
        return re.sub(r'[\x00-\x1F]+', '', text)
    return text

file_name = 'processed_data.xlsx'
try:
    wb = load_workbook(file_name)
    sheet1 = wb.active
    cnt = sheet1.max_row + 1
except FileNotFoundError:
    wb = Workbook()
    sheet1 = wb.active
    headers = ["Id", "Title", "Province", "District", "Ward", "Street", "Frontage", "MainOrientation", 
               "BalconyOrientation", "NumberOfFloor", "NumberOfBedroom", "NumberOfBathroom", 
               "LegalDocumentType", "Furnishing", "Address", "Owner", "Area", "Price", "Description"]
    sheet1.append(headers)
    cnt = 2

driver = uc.Chrome(version_main=130)
pageNumber = 1

for index in range(2, 150):
    pageNumber = index    
    url = f'https://batdongsan.com.vn/ban-nha-rieng-long-bien/p{pageNumber}?cIds=163&disIds=8,15,2'
    print("TRANG MỚI: ", url)
    driver.get(url)
    time.sleep(2.5)
    
    lst = [element.get_attribute("href") for element in driver.find_elements(By.CSS_SELECTOR, "#product-lists-web a")]
    
    for itm in lst:
        if itm is None or itm == "" or "vaymuanha.batdongsan.com.vn" in itm or "batdongsan.com.vn/unknow-page" in itm:
            print("ERROR: ", itm)
            continue
    
        try:
            print("detail: ", pageNumber, itm)
            driver.get(itm)
            time.sleep(1.5)
        except:
            print(itm)
            wb.save(file_name)
    
        title, path_menu, address, province, district, ward, street = None, None, None, None, None, None, None
        description, phone_number, area, price, frontage, mainOrientation = None, None, None, None, None, None
        balconyOrientation, numberOfFloor, numberOfBedroom, numberOfBathroom, legalDocumentType, furnishing = None, None, None, None, None, None
        owner = None

        try:
            title = driver.find_element(By.CSS_SELECTOR, "h1[class*=js__pr-title]").text
        except:
            title = None

        try:
            path_menu = driver.find_element(By.CSS_SELECTOR, "[class*=re__breadcrumb]").text
        except:
            path_menu = None

        try:
            address = driver.find_element(By.CSS_SELECTOR, "span[class*=js__pr-address]").text
            address = re.sub(r'\b(Đường|Phường|Phố)\b\s*', '', address)
            temp = address.split(", ")
            temp.reverse()
            province, district, ward, street = (temp + [None] * 4)[:4]
        except:
            pass

        try:
            description = driver.find_element(By.CSS_SELECTOR, "[class*=re__pr-description] > div").text
        except:
            description = None

        try:
            phone_number = driver.find_element(By.CSS_SELECTOR, "[class*=re__pr-scrollbar-tablet] > a").get_attribute("data-href")
            phone_number = phone_number.replace("sms://", "").split("/")[0]
        except:
            phone_number = None

        try:
            elements = driver.find_elements(By.XPATH, "//div[contains(@class, 're__pr-specs-content-item')]")
            for element in elements:
                label = element.find_element(By.XPATH, ".//span[contains(@class, 're__pr-specs-content-item-title')]").text
                value = element.find_element(By.XPATH, ".//span[contains(@class, 're__pr-specs-content-item-value')]").text

                if label == "Diện tích":
                    area = value.split(" ")[0]
                elif label == "Mức giá":
                    price = convert_price(value, area)
                elif label == "Mặt tiền":
                    if value:
                        # Chuyển đổi dấu phẩy thành dấu chấm nếu cần thiết
                        frontage = convert_comma_to_dot(value.split(" ")[0])
                    else:
                        frontage = None
                elif label == "Hướng nhà":
                    mainOrientation = value
                elif label == "Hướng ban công":
                    balconyOrientation = value
                elif label == "Số tầng":
                    numberOfFloor = value.split(" ")[0]
                elif label == "Số phòng ngủ":
                    numberOfBedroom = value.split(" ")[0]
                elif label == "Số toilet":
                    numberOfBathroom = value.split(" ")[0]
                elif label == "Pháp lý":
                    legalDocumentType = value
                elif label == "Nội thất":
                    furnishing = value
            if frontage is None and description:
                frontage = extract_frontage_from_description(description)
                
            if mainOrientation is None and description:
                mainOrientation = extract_main_orientation(description)
        except:
            pass

        try:
            owner = driver.find_element(By.CSS_SELECTOR, "[class*=re__contact-name] > a").text
        except:
            owner = None
        if title != None or title != "":
            sheet1.append([cnt, remove_illegal_characters(normalize_string(title)), remove_illegal_characters(normalize_string(province)), 
                           remove_illegal_characters(normalize_string(district)), 
                       remove_illegal_characters(normalize_string(ward)), remove_illegal_characters(normalize_string(street)), frontage,
                       remove_illegal_characters(normalize_string(mainOrientation)), remove_illegal_characters(normalize_string(balconyOrientation)),
                       numberOfFloor, numberOfBedroom, numberOfBathroom, remove_illegal_characters(normalize_string(legalDocumentType)), 
                       remove_illegal_characters(normalize_string(furnishing)), remove_illegal_characters(normalize_string(address)), 
                       remove_illegal_characters(normalize_string(owner)), area, price, remove_illegal_characters(normalize_string(description))])
            min = int((time.time() - start_time)//60)
            sec = int((time.time() - start_time)%60)
            print(f'done: {cnt}. Thời gian chạy: {min} phút {sec} giây')
        else: 
            continue
        wb.save(file_name)
        cnt += 1

wb.save(file_name)

driver.close()

end_time = time.time()
elapsed_time = end_time - start_time
minutes = int(elapsed_time // 60)
seconds = int(elapsed_time % 60)

print(f"Thời gian chạy: {minutes} phút {seconds} giây")